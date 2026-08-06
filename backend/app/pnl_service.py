"""
Shared P&L aggregation for the /api/paper/pnl/* endpoints (summary + Excel export): per-strategy
and combined totals, a daily net-P&L series, and the raw closed-trade log — all DB-backed
(options_positions), scoped to a [start_date, end_date] range (see src/utils/date_ranges.py).

net P&L = gross realized_pnl minus both legs' charges (entry_charges + exit_charges) — the same
"actual cash-flow effect on the wallet" meaning as Order.net_pnl in src/simulator/paper_trader.py,
just aggregated from Postgres rows instead of in-memory Order objects.
"""
from datetime import date

from openpyxl import Workbook

from . import db


async def strategy_pnl_rows(start_date: date, end_date: date) -> dict[str, dict]:
    """{strategy: {trades, wins, gross_pnl, charges}} for trades that closed within range."""
    try:
        pool = db.get_pool()
    except RuntimeError:
        return {}
    rows = await pool.fetch(
        """SELECT strategy, COUNT(*) AS trades,
                  SUM(CASE WHEN realized_pnl > 0 THEN 1 ELSE 0 END) AS wins,
                  COALESCE(SUM(realized_pnl), 0) AS gross_pnl,
                  COALESCE(SUM(entry_charges + exit_charges), 0) AS charges
           FROM options_positions
           WHERE status = 'CLOSED' AND strategy IS NOT NULL AND exit_time::date BETWEEN $1 AND $2
           GROUP BY strategy""",
        start_date, end_date,
    )
    return {row["strategy"]: dict(row) for row in rows}


async def daily_net_pnl_series(start_date: date, end_date: date) -> list[dict]:
    try:
        pool = db.get_pool()
    except RuntimeError:
        return []
    rows = await pool.fetch(
        """SELECT exit_time::date AS day,
                  COALESCE(SUM(realized_pnl - entry_charges - exit_charges), 0) AS net_pnl
           FROM options_positions
           WHERE status = 'CLOSED' AND exit_time::date BETWEEN $1 AND $2
           GROUP BY exit_time::date ORDER BY day""",
        start_date, end_date,
    )
    return [{"date": row["day"].isoformat(), "net_pnl": float(row["net_pnl"])} for row in rows]


async def closed_trades_in_range(start_date: date, end_date: date) -> list[dict]:
    try:
        pool = db.get_pool()
    except RuntimeError:
        return []
    rows = await pool.fetch(
        """SELECT order_id, strategy, symbol, qty, entry_price, entry_time, exit_price, exit_time,
                  exit_reason, realized_pnl, entry_charges, exit_charges
           FROM options_positions
           WHERE status = 'CLOSED' AND exit_time::date BETWEEN $1 AND $2
           ORDER BY exit_time""",
        start_date, end_date,
    )
    return [dict(row) for row in rows]


def build_strategy_summary(strategy_names: list[str], db_rows: dict[str, dict],
                            wallets: dict[str, dict]) -> list[dict]:
    """One row per known strategy, even with zero trades in range — the Individual P&L view
    needs a stable strategy list, not just whichever ones happened to trade."""
    result = []
    for name in strategy_names:
        row = db_rows.get(name, {})
        trades = row.get("trades") or 0
        wins = row.get("wins") or 0
        gross_pnl = float(row.get("gross_pnl") or 0)
        charges = float(row.get("charges") or 0)
        wallet = wallets.get(name)
        result.append({
            "strategy": name, "trades": trades,
            "win_rate": round(wins / trades * 100, 2) if trades else 0.0,
            "gross_pnl": round(gross_pnl, 2), "charges": round(charges, 2),
            "net_pnl": round(gross_pnl - charges, 2),
            "wallet_balance": wallet["balance"] if wallet else None,
            "allocated_capital": wallet["allocated_capital"] if wallet else None,
        })
    return result


def combine_totals(strategy_rows: list[dict]) -> dict:
    wallet_values = [s["wallet_balance"] for s in strategy_rows if s["wallet_balance"] is not None]
    allocated_values = [s["allocated_capital"] for s in strategy_rows if s["allocated_capital"] is not None]
    return {
        "trades": sum(s["trades"] for s in strategy_rows),
        "gross_pnl": round(sum(s["gross_pnl"] for s in strategy_rows), 2),
        "charges": round(sum(s["charges"] for s in strategy_rows), 2),
        "net_pnl": round(sum(s["net_pnl"] for s in strategy_rows), 2),
        "wallet_balance": round(sum(wallet_values), 2) if wallet_values else None,
        "allocated_capital": round(sum(allocated_values), 2) if allocated_values else None,
    }


def build_workbook(strategies: list[dict], combined: dict, trades: list[dict],
                    start_date: date, end_date: date) -> Workbook:
    wb = Workbook()

    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Range", f"{start_date.isoformat()} to {end_date.isoformat()}"])
    summary_sheet.append([])
    summary_sheet.append(
        ["Total Trades", "Gross P&L", "Charges", "Net P&L", "Wallet Balance", "Allocated Capital"])
    summary_sheet.append([
        combined["trades"], combined["gross_pnl"], combined["charges"], combined["net_pnl"],
        combined["wallet_balance"], combined["allocated_capital"],
    ])

    by_strategy_sheet = wb.create_sheet("By Strategy")
    by_strategy_sheet.append(
        ["Strategy", "Trades", "Win Rate %", "Gross P&L", "Charges", "Net P&L",
         "Wallet Balance", "Allocated Capital"])
    for s in strategies:
        by_strategy_sheet.append([
            s["strategy"], s["trades"], s["win_rate"], s["gross_pnl"], s["charges"], s["net_pnl"],
            s["wallet_balance"], s["allocated_capital"],
        ])

    trades_sheet = wb.create_sheet("Trades")
    trades_sheet.append(
        ["Order ID", "Strategy", "Symbol", "Qty", "Entry Time", "Entry Price", "Exit Time",
         "Exit Price", "Exit Reason", "Gross P&L", "Charges", "Net P&L"])
    for t in trades:
        gross = float(t["realized_pnl"]) if t["realized_pnl"] is not None else None
        charges = float(t["entry_charges"] or 0) + float(t["exit_charges"] or 0)
        net = round(gross - charges, 2) if gross is not None else None
        # openpyxl can't write tz-aware datetimes -- strip tzinfo (values are for a human export,
        # not further computation).
        entry_time = t["entry_time"].replace(tzinfo=None) if t["entry_time"] else None
        exit_time = t["exit_time"].replace(tzinfo=None) if t["exit_time"] else None
        trades_sheet.append([
            t["order_id"], t["strategy"], t["symbol"], t["qty"], entry_time,
            float(t["entry_price"]) if t["entry_price"] is not None else None, exit_time,
            float(t["exit_price"]) if t["exit_price"] is not None else None,
            t["exit_reason"], gross, round(charges, 2), net,
        ])

    return wb
